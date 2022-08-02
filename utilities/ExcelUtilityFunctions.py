import openpyxl
from openpyxl.styles import PatternFill    # for colouring

def getRowCount(fileName,sheetName):
    workbook=openpyxl.load_workbook(fileName)
    sheet=workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(fileName,sheetName):
    workbook=openpyxl.load_workbook(fileName)
    sheet=workbook[sheetName]
    return(sheet.max_column)

def readData(fileName,sheetName,rownum,columnnum):
    workbook=openpyxl.load_workbook(fileName)
    sheet=workbook[sheetName]
    return sheet.cell(rownum,columnnum).value

def writeData(fileName,sheetName,rownum,columnnum,data):
    workbook=openpyxl.load_workbook(fileName)
    sheet=workbook[sheetName]
    sheet.cell(rownum,columnnum).value=data
    workbook.save(fileName)

def fillGreenColor(fileName,sheetName,rownum,columnnum):
    workbook=openpyxl.load_workbook(fileName)
    sheet=workbook[sheetName]
    greenFill=PatternFill(start_color='60b212',end_color='60b212',fill_type='solid')
    sheet.cell(rownum,columnnum).fill=greenFill
    workbook.save(fileName)

def fillRedColor(fileName,sheetName,rownum,columnnum):
    workbook=openpyxl.load_workbook(fileName)
    sheet=workbook[sheetName]
    redFill=PatternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')
    sheet.cell(rownum,columnnum).fill=redFill
    workbook.save(fileName)